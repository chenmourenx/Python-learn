"""
需求：将工作表1中price乘0.9，并保存到后一列中
    然后将price绘制乘图表，最后保存表
"""

import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_work(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']  # 指定工作表
    # cell = sheet['a1']  # 指定单元格，方法一
    # cell = sheet.cell(1, 1)  # 指定单元格，方法二
    # print(cell.value)  # 单元格值
    # print(sheet.max_row) # 行数

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # print(cell.value) price的值
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # 定义数据范围
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # 绘制图表到e2
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


if __name__ == '__main__':
    filename = input('表')
    process_work(filename)
